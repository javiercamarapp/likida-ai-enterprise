# -*- coding: utf-8 -*-
"""test_data_migration.py — Tests del módulo de migración de datos.

Cubre:
  - Validación de RFC mexicano (personas físicas/morales, genérico, inválido).
  - Importador de Excel (.xlsx) — Clientes, CFDIs, Cuentas, Empleados.
  - Importador de CSV — detección de tipo + normalización.
  - Mapeo CONTPAQi al esquema canónico.
  - MigrationService: start/validate/execute/status.
  - Endpoints API /api/v1/migration/* (upload/execute/status/errors).

Nota: los .xlsx de prueba se generan en memoria con openpyxl.
"""
import csv
import io
import os

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from b2b_ai.features.data_migration import models as dm_models
from b2b_ai.features.data_migration.importers.contpaqi_mapper import ContpaqiMapper
from b2b_ai.features.data_migration.importers.csv_importer import ImportCSVData
from b2b_ai.features.data_migration.importers.excel_importer import (
    ExcelImportError,
    ImportClientData,
)
from b2b_ai.features.data_migration.models import (
    MigrationDataType,
    MigrationFileType,
    MigrationStatus,
)
from b2b_ai.features.data_migration.routes import build_data_migration_router
from b2b_ai.features.data_migration.service import MigrationError, MigrationService
from b2b_ai.features.data_migration.validators.rfc_validator import (
    describe_rfc_error,
    is_valid_mx_rfc,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(autouse=True)
def _reset():
    dm_models._reset_state()
    yield
    dm_models._reset_state()


@pytest.fixture
def client():
    """TestClient con el router de migración (auth stub)."""
    app = FastAPI()

    def fake_require_api_key():
        return {"tenant_id": "tenant_test_123", "api_key": "key"}

    app.include_router(
        build_data_migration_router(db=None, require_api_key=fake_require_api_key)
    )
    return TestClient(app)


def _make_xlsx(tmp_path, name="despacho.xlsx"):
    """Crea un .xlsx de prueba con hojas de clientes, cfdi y empleados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["RFC", "Razon Social", "Regimen Fiscal"])
    ws.append(["GAGF8607255K5", "Grupo Andrés Galván, S.A. de C.V.", "601"])
    ws.append(["XAXX010101000", "Cliente Genérico", "616"])
    ws.append(["RFCINVALIDO", "Cliente Malo", "601"])  # RFC inválido

    ws2 = wb.create_sheet("CFDIs")
    ws2.append(["UUID", "Emisor RFC", "Receptor RFC", "Total", "Fecha", "Concepto"])
    ws2.append(["A1B2C3D4-E5F6-4A5B-8C9D-0123456789AB", "GAGF8607255K5",
                "XAXX010101000", 1500.50, "2026-01-15", "Honorarios"])
    ws2.append(["00000000-0000-0000-0000-000000000000", "GAGF8607255K5",
                "RFC MALO", 200, "2026-01-16", "Sin RFC"])  # receptor inválido

    ws3 = wb.create_sheet("Empleados")
    ws3.append(["RFC", "Nombre", "Salario", "Puesto"])
    ws3.append(["HELO0205249S4", "Héctor López Ortega", 25000, "Contador Sr."])
    ws3.append(["", "Empleado Sin RFC", 15000, "Auxiliar"])  # RFC opcional

    path = tmp_path / name
    wb.save(path)
    return str(path)


def _make_csv(tmp_path, name="clientes.csv"):
    """Crea un CSV de prueba de clientes (nombre de archivo => tipo cliente)."""
    path = tmp_path / name
    with open(path, "w", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["RFC", "Razon Social", "Regimen Fiscal"])
        w.writerow(["GAGF8607255K5", "Grupo Andrés Galván, S.A. de C.V.", "601"])
        w.writerow(["XAXX010101000", "Cliente Genérico", "616"])
    return str(path)


# ---------------------------------------------------------------------------
# RFC validator
# ---------------------------------------------------------------------------

class TestRFCValidator:
    def test_rfc_generico(self):
        assert is_valid_mx_rfc("XAXX010101000")
        assert is_valid_mx_rfc("XEXX010101000")

    def test_rfc_moral_valido(self):
        assert is_valid_mx_rfc("GAGF8607255K5")

    def test_rfc_fisica_valido(self):
        assert is_valid_mx_rfc("HELO0205249S4")

    def test_rfc_normaliza(self):
        assert is_valid_mx_rfc(" gagf8607255k5 ")

    def test_rfc_invalido(self):
        assert not is_valid_mx_rfc("")
        assert not is_valid_mx_rfc("ABC123")
        assert not is_valid_mx_rfc("INVALIDO00000")
        assert not is_valid_mx_rfc("123456789012")
        # Fecha imposible (mes 99)
        assert not is_valid_mx_rfc("GAGF9999995K5")

    def test_describe_error(self):
        assert "RFC vacío" in describe_rfc_error("")
        assert describe_rfc_error("GAGF8607255K5") == ""


# ---------------------------------------------------------------------------
# Importador Excel
# ---------------------------------------------------------------------------

class TestExcelImporter:
    def test_parse_clientes_y_tipos(self, tmp_path):
        path = _make_xlsx(tmp_path)
        items = ImportClientData().parse_excel(path)
        tipos = sorted({it.data_type.value for it in items})
        assert "cliente" in tipos
        assert "cfdi" in tipos
        assert "empleado" in tipos

        clientes = [it for it in items if it.data_type == MigrationDataType.CLIENTE]
        assert len(clientes) == 3
        assert clientes[0].data["rfc"] == "GAGF8607255K5"
        assert clientes[0].data["razon_social"]
        assert clientes[0].data["regimen_fiscal"] == "601"

    def test_parse_cfdi(self, tmp_path):
        path = _make_xlsx(tmp_path)
        items = ImportClientData().parse_excel(path)
        cfdis = [it for it in items if it.data_type == MigrationDataType.CFDI]
        assert len(cfdis) == 2
        assert cfdis[0].data["total"] == 1500.50
        assert cfdis[0].data["emisor_rfc"] == "GAGF8607255K5"

    def test_archivo_inexistente(self):
        with pytest.raises(ExcelImportError):
            ImportClientData().parse_excel("/no/existe.xlsx")


# ---------------------------------------------------------------------------
# Importador CSV
# ---------------------------------------------------------------------------

class TestCSVImporter:
    def test_inferencia_por_nombre(self, tmp_path):
        path = _make_csv(tmp_path)
        items = ImportCSVData().parse_csv(path)
        assert items
        assert all(it.data_type == MigrationDataType.CLIENTE for it in items)
        assert items[0].data["rfc"] == "GAGF8607255K5"

    def test_type_forzado(self, tmp_path):
        path = _make_csv(tmp_path, name="datos.csv")
        items = ImportCSVData(data_type=MigrationDataType.CLIENTE).parse_csv(path)
        assert items and items[0].data["rfc"] == "GAGF8607255K5"

    def test_sin_tipo_inferible(self, tmp_path):
        p = tmp_path / "datos.csv"
        p.write_text("RFC,Razon Social\nGAGF8607255K5,Grupo\n", encoding="utf-8")
        with pytest.raises(Exception):
            ImportCSVData().parse_csv(str(p))


# ---------------------------------------------------------------------------
# Mapeo CONTPAQi
# ---------------------------------------------------------------------------

class TestContpaqiMapper:
    def test_mapea_clientes(self):
        mapper = ContpaqiMapper()
        rows = [
            {"RFC": "GAGF8607255K5", "Nombre": "Grupo Andrés", "Regimen Fiscal": "601"},
            {"RFC": "XAXX010101000", "Nombre": "Genérico", "Regimen Fiscal": "616"},
        ]
        items = mapper.map_sheet("Clientes", rows)
        assert len(items) == 2
        assert items[0].data["rfc"] == "GAGF8607255K5"
        assert items[0].data["razon_social"] == "Grupo Andrés"
        assert items[0].data["regimen_fiscal"] == "601"

    def test_inferencia_por_columnas(self):
        mapper = ContpaqiMapper()
        rows = [{"RFC": "GAGF8607255K5", "Nombre": "X"}]
        assert mapper.infer_sheet_type("Hoja1", rows) == MigrationDataType.CLIENTE

    def test_hoja_desconocida(self):
        mapper = ContpaqiMapper()
        assert mapper.map_sheet("Otros", [{"a": "1"}]) == []


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class TestMigrationService:
    def test_start_migration_excel(self, tmp_path):
        svc = MigrationService()
        path = _make_xlsx(tmp_path)
        job = svc.start_migration(path, "excel", "tenant_1")
        assert job.status == MigrationStatus.VALIDATED
        assert job.total_items == 7  # 3 clientes + 2 cfdi + 2 empleados
        assert job.valid_count >= 1
        assert job.invalid_count >= 1  # RFCINVALIDO + receptor RFC MALO

    def test_start_migration_csv(self, tmp_path):
        svc = MigrationService()
        path = _make_csv(tmp_path)
        job = svc.start_migration(path, "csv", "tenant_1")
        assert job.status == MigrationStatus.VALIDATED
        assert job.total_items == 2
        assert job.valid_count == 2
        assert job.invalid_count == 0

    def test_validate_data(self):
        svc = MigrationService()
        item = dm_models.MigrationItem(
            data_type=MigrationDataType.CLIENTE,
            data={"rfc": "GAGF8607255K5", "razon_social": "Grupo X"},
        )
        res = svc.validate_data([item])
        assert res["valid_count"] == 1
        assert item.valid is True

    def test_execute_migration(self, tmp_path):
        svc = MigrationService()
        path = _make_csv(tmp_path)
        job = svc.start_migration(path, "csv", "tenant_1")
        job = svc.execute_migration(job.id)
        assert job.status == MigrationStatus.COMPLETED
        assert job.imported_count == 2
        assert job.failed_count == 0

    def test_execute_migration_partial(self, tmp_path):
        svc = MigrationService()
        path = _make_xlsx(tmp_path)
        job = svc.start_migration(path, "excel", "tenant_1")
        job = svc.execute_migration(job.id)
        # Hay 2 inválidos (RFCINVALIDO, RFC MALO) => parcial
        assert job.status == MigrationStatus.PARTIAL
        assert job.failed_count == job.invalid_count
        assert job.imported_count == job.valid_count

    def test_execute_desconocido(self):
        svc = MigrationService()
        with pytest.raises(MigrationError):
            svc.execute_migration("no-existe")

    def test_get_status(self, tmp_path):
        svc = MigrationService()
        path = _make_csv(tmp_path)
        job = svc.start_migration(path, "csv", "tenant_1")
        assert svc.get_migration_status(job.id) is not None
        assert svc.get_migration_status("zzz") is None

    def test_archivo_inexistente(self):
        svc = MigrationService()
        with pytest.raises(MigrationError):
            svc.start_migration("/no/existe.csv", "csv", "tenant_1")

    def test_file_type_invalido(self, tmp_path):
        svc = MigrationService()
        path = _make_csv(tmp_path)
        with pytest.raises(MigrationError):
            svc.start_migration(path, "pdf", "tenant_1")


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

class TestMigrationAPI:
    def test_upload_y_ciclo(self, client, tmp_path):
        path = _make_csv(tmp_path)
        with open(path, "rb") as f:
            r = client.post(
                "/api/v1/migration/upload",
                files={"file": ("clientes.csv", f, "text/csv")},
                params={"file_type": "csv"},
            )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["ok"] is True
        job_id = body["data"]["job_id"]

        # status
        r = client.get(f"/api/v1/migration/{job_id}/status")
        assert r.status_code == 200
        assert r.json()["job"]["status"] == "validated"

        # errors (todos válidos => vacío)
        r = client.get(f"/api/v1/migration/{job_id}/errors")
        assert r.status_code == 200
        assert r.json()["errors"] == []

        # execute
        r = client.post(f"/api/v1/migration/{job_id}/execute")
        assert r.status_code == 200
        assert r.json()["data"]["status"] == "completed"

    def test_upload_extension_invalida(self, client, tmp_path):
        p = tmp_path / "datos.txt"
        p.write_text("RFC\nGAGF8607255K5\n", encoding="utf-8")
        with open(p, "rb") as f:
            r = client.post(
                "/api/v1/migration/upload",
                files={"file": ("datos.txt", f, "text/plain")},
            )
        assert r.status_code == 400

    def test_status_404(self, client):
        r = client.get("/api/v1/migration/inexistente/status")
        assert r.status_code == 404

    def test_errors_son_visibles(self, client, tmp_path):
        path = _make_xlsx(tmp_path)
        with open(path, "rb") as f:
            r = client.post(
                "/api/v1/migration/upload",
                files={"file": ("despacho.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                params={"file_type": "excel"},
            )
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        r = client.get(f"/api/v1/migration/{job_id}/errors")
        assert r.status_code == 200
        errors = r.json()["errors"]
        assert len(errors) >= 1  # RFCINVALIDO y/o RFC MALO
