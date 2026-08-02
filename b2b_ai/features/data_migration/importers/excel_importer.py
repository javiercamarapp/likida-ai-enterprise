# -*- coding: utf-8 -*-
"""excel_importer.py — Importación de datos desde archivos Excel (.xlsx).

`ImportClientData` parsea un libro de Excel exportado de CONTPAQi (o generado
a mano) y extrae ítems normalizados de:

    - Clientes (RFC, razón social, régimen fiscal)
    - CFDIs (UUID, emisor, receptor, total, fecha, conceptos)
    - Cuentas bancarias (número, banco, saldo)
    - Empleados (RFC, nombre, salario, puesto)

Cada hoja se interpreta según su nombre (normalizado). Se soportan también
columnas con encabezados variantes mediante alias (ver importers/base.py).
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from b2b_ai.features.data_migration.importers.base import (
    detect_sheet_type,
    normalize_row,
)
from b2b_ai.features.data_migration.models import (
    MigrationDataType,
    MigrationItem,
)

logger = logging.getLogger("b2b_ai.data_migration")

MAX_ROWS = 100_000


class ExcelImportError(Exception):
    """Error al abrir o parsear el archivo Excel."""


class ImportClientData:
    """Parsa un libro Excel y produce ítems de migración normalizados."""

    def __init__(self, sheet_types: Optional[Dict[str, MigrationDataType]] = None):
        self.sheet_types = sheet_types or {}

    # ------------------------------------------------------------------
    def parse_excel(self, path) -> List[MigrationItem]:
        """Abre el .xlsx y devuelve los ítems de todas las hojas reconocidas."""
        file_path = Path(path)
        if not file_path.exists():
            raise ExcelImportError(f"El archivo no existe: {file_path}")
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ExcelImportError(f"No se pudo abrir el Excel: {exc}") from exc

        items: List[MigrationItem] = []
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data_type = self._resolve_type(sheet_name)
                if data_type is None:
                    logger.info("hoja ignorada (tipo desconocido): %s", sheet_name)
                    continue
                items.extend(self._parse_sheet(ws, sheet_name, data_type))
        finally:
            wb.close()
        return items

    def _resolve_type(self, sheet_name: str) -> Optional[MigrationDataType]:
        if sheet_name in self.sheet_types:
            return self.sheet_types[sheet_name]
        return detect_sheet_type(sheet_name)

    def _parse_sheet(self, ws, sheet_name: str,
                     data_type: MigrationDataType) -> List[MigrationItem]:
        """Convierte las filas de una hoja a ítems."""
        header: Optional[List[Any]] = None
        items: List[MigrationItem] = []
        row_idx = 0
        for row in ws.iter_rows(values_only=True):
            row_idx += 1
            if row_idx > MAX_ROWS:
                logger.warning("hoja %s supera %d filas; truncando",
                               sheet_name, MAX_ROWS)
                break
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            if header is None:
                header = [str(c) if c is not None else "" for c in row]
                continue
            record: Dict[str, Any] = {}
            for col_name, value in zip(header, row):
                record[col_name] = value
            if not any(str(v).strip() != "" for v in record.values()):
                continue
            data = normalize_row(record, data_type)
            items.append(MigrationItem(
                data_type=data_type,
                source=sheet_name,
                row=row_idx,
                data=data,
            ))
        return items
