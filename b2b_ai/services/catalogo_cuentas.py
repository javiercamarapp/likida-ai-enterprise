# -*- coding: utf-8 -*-
"""
catalogo_cuentas.py — Catálogo de cuentas (CUC) conforme a la Norma Técnica
2026 del SAT (Código Agrupador) para contabilidad electrónica.

Clase principal:
    CatalogoCuentas — gestiona un catálogo de cuentas del SAT.

Responsabilidades:
  - Catálogo por defecto (estructura tipo SAT, nivel + naturaleza).
  - Validación estructural de cada cuenta (codigo, descripcion, nivel,
    naturaleza deudora/acreedora).
  - Generación del XML del catálogo para envío al SAT.
  - Importación del catálogo desde CSV / Excel del cliente.
  - Asignación automática de cuentas del catálogo a las clasificaciones
    del agente (categorías de factura -> cuenta).

> La presentación oficial exige e.firma / software del SAT. Este módulo
> prepara y valida el archivo; no hace el envío real.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# Naturalezas válidas del SAT: D = deudora, A = acreedora.
NATURALEZAS_VALIDAS: Tuple[str, ...] = ("D", "A")


@dataclass
class Cuenta:
    """Una cuenta del catálogo. Naturaleza: 'D' (deudora) o 'A' (acreedora)."""
    codigo: str
    descripcion: str
    nivel: int
    naturaleza: str
    grupo: str = ""

    def as_dict(self) -> Dict[str, Any]:
        return {
            "codigo": self.codigo,
            "descripcion": self.descripcion,
            "nivel": self.nivel,
            "naturaleza": self.naturaleza,
            "grupo": self.grupo,
        }


# Catálogo por defecto (Código Agrupador SAT / NT 2026, simplificado).
# Formato: (codigo, descripcion, nivel, naturaleza, grupo)
_CATALOGO_BASE = [
    ("1000", "ACTIVO", 1, "D", "ACTIVO"),
    ("1100", "ACTIVO CIRCULANTE", 2, "D", "ACTIVO"),
    ("1101", "BANCOS", 3, "D", "ACTIVO"),
    ("1102", "CLIENTES", 3, "D", "ACTIVO"),
    ("1103", "IVA ACREDITABLE", 3, "D", "ACTIVO"),
    ("1104", "INVENTARIOS", 3, "D", "ACTIVO"),
    ("1200", "ACTIVO FIJO", 2, "D", "ACTIVO"),
    ("1201", "MOBILIARIO Y EQUIPO", 3, "D", "ACTIVO"),
    ("1202", "EQUIPO DE COMPUTO", 3, "D", "ACTIVO"),
    ("1300", "ACTIVOS DIFERIDOS", 2, "D", "ACTIVO"),
    ("2000", "PASIVO", 1, "A", "PASIVO"),
    ("2100", "PASIVO CIRCULANTE", 2, "A", "PASIVO"),
    ("2101", "PROVEEDORES", 3, "A", "PASIVO"),
    ("2102", "ACREEDORES DIVERSOS", 3, "A", "PASIVO"),
    ("2103", "IVA POR ACREDITAR", 3, "A", "PASIVO"),
    ("2104", "IMPUESTOS POR PAGAR", 3, "A", "PASIVO"),
    ("2200", "PASIVO A LARGO PLAZO", 2, "A", "PASIVO"),
    ("3000", "CAPITAL CONTABLE", 1, "A", "CAPITAL"),
    ("3100", "CAPITAL SOCIAL", 2, "A", "CAPITAL"),
    ("3200", "RESULTADOS ACUMULADOS", 2, "A", "CAPITAL"),
    ("4000", "INGRESOS", 1, "A", "INGRESOS"),
    ("4100", "INGRESOS POR SERVICIOS", 2, "A", "INGRESOS"),
    ("4200", "OTROS INGRESOS", 2, "A", "INGRESOS"),
    ("5000", "COSTOS", 1, "D", "COSTOS"),
    ("5100", "COSTO DE VENTAS", 2, "D", "COSTOS"),
    ("6000", "GASTOS", 1, "D", "GASTOS"),
    ("6100", "GASTOS DE OPERACION", 2, "D", "GASTOS"),
    ("6101", "SUELDOS Y SALARIOS", 3, "D", "GASTOS"),
    ("6102", "GASTOS ADMINISTRATIVOS", 3, "D", "GASTOS"),
    ("6103", "GASTOS FINANCIEROS", 3, "D", "GASTOS"),
    ("6200", "GASTOS NO DEDUCIBLES", 2, "D", "GASTOS"),
]

# Mapeo categoría de factura del agente -> cuenta contable por defecto.
CATEGORIA_A_CUENTA = {
    "gasto_operativo": "6102",
    "inversion": "1202",
    "activo_fijo": "1201",
    "nomina": "6101",
    "ingreso": "4100",
    "venta": "4100",
    "servicios": "4100",
    "costo": "5100",
    "impuestos": "2104",
    "desconocido": "6102",
}


class CatalogoCuentas:
    """Catálogo de cuentas contables del SAT.

    Uso:
        cat = CatalogoCuentas()
        cat.validar()                      # lanza si hay cuentas mal formadas
        xml = cat.generar_xml(rfc, 2026, 7)
        cat.importar_csv("catalogo.csv")
    """

    def __init__(self, cuentas: Optional[List[Any]] = None) -> None:
        if cuentas is None:
            self.cuentas: List[Cuenta] = [Cuenta(c, d, n, t, g) for c, d, n, t, g in _CATALOGO_BASE]
        else:
            self.cuentas = [Cuenta(**c) if isinstance(c, dict) else c for c in cuentas]
        self._index: Optional[Dict[str, Cuenta]] = None
        self._reindex()

    # ---- Construcción / índice -------------------------------------------
    def _reindex(self) -> None:
        self._index = {c.codigo: c for c in self.cuentas}

    def add(self, codigo: str, descripcion: str, nivel: int,
            naturaleza: str, grupo: str = "") -> Cuenta:
        """Añade una cuenta. Devuelve la cuenta creada."""
        cuenta = Cuenta(str(codigo), descripcion, int(nivel),
                        naturaleza.upper(), grupo)
        self.cuentas.append(cuenta)
        self._reindex()
        return cuenta

    def find(self, codigo: str) -> Optional[Cuenta]:
        return self._index.get(str(codigo))

    def __len__(self) -> int:
        return len(self.cuentas)

    # ---- Validación --------------------------------------------------------
    def errores(self) -> List[str]:
        """Devuelve la lista de errores de validación del catálogo.

        Cada cuenta debe tener: código, descripción, nivel (>=1) y
        naturaleza deudora ('D') o acreedora ('A'). Detecta además
        códigos duplicados.
        """
        errs = []
        seen = set()
        for c in self.cuentas:
            if not c.codigo or not str(c.codigo).strip():
                errs.append("Cuenta sin código.")
                continue
            if str(c.codigo) in seen:
                errs.append("Código duplicado: %s" % c.codigo)
            seen.add(str(c.codigo))
            if not c.descripcion or not str(c.descripcion).strip():
                errs.append("Cuenta %s sin descripción." % c.codigo)
            try:
                nivel = int(c.nivel)
            except (TypeError, ValueError):
                errs.append("Cuenta %s con nivel inválido: %r" % (c.codigo, c.nivel))
                nivel = 0
            if nivel < 1:
                errs.append("Cuenta %s con nivel < 1." % c.codigo)
            if str(c.naturaleza).upper() not in NATURALEZAS_VALIDAS:
                errs.append("Cuenta %s con naturaleza inválida: %r"
                            % (c.codigo, c.naturaleza))
        return errs

    def validar(self) -> bool:
        """Valida el catálogo completo. Devuelve True si es válido.

        Lanza ValueError con el detalle si hay errores."""
        errs = self.errores()
        if errs:
            raise ValueError("Catálogo inválido: " + "; ".join(errs))
        return True

    # ---- Importación ------------------------------------------------------
    def importar_csv(self, ruta: str) -> int:
        """Importa un catálogo desde CSV del cliente. Devuelve cuentas leídas.

        Se aceptan las columnas (en cualquier orden o por cabecera):
            codigo, descripcion, nivel, naturaleza[, grupo]
        Sin cabecera, se asume el orden: codigo,descripcion,nivel,naturaleza.
        Reemplaza el catálogo actual.
        """
        import csv

        with open(ruta, newline="", encoding="utf-8-sig") as fh:
            sample = fh.read(2048)
            fh.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            rows = list(csv.reader(fh, dialect))
        if not rows:
            raise ValueError("CSV vacío: %s" % ruta)
        header = [h.strip().lower() for h in rows[0]]
        if any(h in header for h in ("codigo", "cuenta", "nivel")):
            rows = rows[1:]
            col = {k: header.index(k) for k in header
                   if k in ("codigo", "cuenta", "descripcion", "desc", "nivel",
                            "naturaleza", "tipo", "grupo")}
            def _extract(r, *names):
                for n in names:
                    if n in col:
                        return r[col[n]]
                return ""
        else:
            def _extract(r, *names):
                idx = {"codigo": 0, "descripcion": 1, "nivel": 2,
                       "naturaleza": 3, "grupo": 4}.get(names[0], 0)
                return r[idx] if idx < len(r) else ""

        nuevas = []
        for r in rows:
            if not r or not r[0].strip():
                continue
            codigo = _extract(r, "codigo", "cuenta").strip()
            desc = _extract(r, "descripcion", "desc").strip()
            nivel = _extract(r, "nivel").strip() or "3"
            natura = _extract(r, "naturaleza", "tipo").strip().upper() or "D"
            grupo = _extract(r, "grupo").strip()
            nuevas.append(Cuenta(codigo, desc, int(nivel), natura, grupo))
        if not nuevas:
            raise ValueError("No se encontraron cuentas en el CSV: %s" % ruta)
        self.cuentas = nuevas
        self._reindex()
        return len(nuevas)

    def importar_excel(self, ruta: str) -> int:
        """Importa el catálogo desde un archivo .xlsx del cliente.

        Lee la primera hoja con cabeceras: codigo, descripcion, nivel,
        naturaleza[, grupo]. Requiere openpyxl. Devuelve cuentas leídas."""
        try:
            from openpyxl import load_workbook
        except ImportError:  # pragma: no cover
            raise ImportError("openpyxl no instalado. pip install openpyxl")
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        header = [str(c.value).strip().lower() if c.value else ""
                  for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {}
        for i, h in enumerate(header):
            if h in ("codigo", "cuenta", "descripcion", "desc", "nivel",
                     "naturaleza", "tipo", "grupo"):
                idx.setdefault(h, i)
        nuevas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            def get(*names):
                for n in names:
                    if n in idx and idx[n] < len(row):
                        return row[idx[n]]
                return ""
            codigo = str(get("codigo", "cuenta") or "").strip()
            if not codigo:
                continue
            desc = str(get("descripcion", "desc") or "").strip()
            nivel = str(get("nivel") or "3").strip() or "3"
            natura = str(get("naturaleza", "tipo") or "D").strip().upper() or "D"
            grupo = str(get("grupo") or "").strip()
            nuevas.append(Cuenta(codigo, desc, int(nivel), natura, grupo))
        if not nuevas:
            raise ValueError("No se encontraron cuentas en: %s" % ruta)
        self.cuentas = nuevas
        self._reindex()
        return len(nuevas)

    def importar_archivo(self, ruta: str) -> int:
        """Detecta el formato (csv/xlsx) y lo importa. Devuelve cuentas leídas."""
        low = ruta.lower()
        if low.endswith((".xlsx", ".xlsm")):
            return self.importar_excel(ruta)
        return self.importar_csv(ruta)

    # ---- Asignación automática ---------------------------------------------
    def asignar_automatico(self, clasificaciones: Optional[Dict[str, Any]]) -> Dict[str, str]:
        """Asigna cuentas del catálogo a las clasificaciones del agente.

        `clasificaciones` es un dict {categoria: cuenta_deseada_o_None}.
        Devuelve {categoria: codigo_cuenta_asignado}, validando que la
        cuenta exista en el catálogo; usa el mapeo por defecto cuando la
        clasificación no especifica cuenta o la cuenta no existe.
        """
        asignacion = {}
        for cat, cuenta in (clasificaciones or {}).items():
            key = str(cat).lower()
            if not cuenta:
                asignacion[key] = CATEGORIA_A_CUENTA.get(key, "6102")
            elif self.find(cuenta):
                asignacion[key] = str(cuenta)
            else:
                asignacion[key] = CATEGORIA_A_CUENTA.get(key, "6102")
        return asignacion

    # ---- Generación XML -----------------------------------------------------
    def generar_xml(self, rfc: str = "", ejercicio: Optional[int] = None,
                    mes: Optional[int] = None) -> str:
        """Genera el XML del catálogo de cuentas conforme al XSD SAT
        (CatalogoCuentas 1.3). Devuelve el XML como string."""
        import xml.sax.saxutils as sx

        self.validar()
        ejercicio = ejercicio or datetime.now().year
        mes = int(mes or 1)
        mes_s = str(mes).zfill(2)
        hoy = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

        def esc(v):
            return sx.escape(str(v or ""))

        ctas = "\n".join(
            '      <Cat:Cta NumCta="%s" Desc="%s" Nivel="%d" Natur="%s"/>'
            % (esc(c.codigo), esc(c.descripcion), int(c.nivel),
               esc(c.naturaleza.upper()))
            for c in sorted(self.cuentas, key=lambda x: x.codigo))
        return (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<Cat:Catalogo '
            'xmlns:Cat="http://www.sat.gob.mx/esquemas/ContabilidadE/1_3/CatalogoCuentas" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
            'xsi:schemaLocation="http://www.sat.gob.mx/esquemas/ContabilidadE/1_3/CatalogoCuentas '
            'http://www.sat.gob.mx/esquemas/ContabilidadE/1_3/CatalogoCuentas/CatalogoCuentas_1_3.xsd" '
            'Version="1.3" RFC="%s" Anio="%s" Mes="%s" '
            'FechaModificacion="%s" Sello="" noCertificado="" Certificado="">\n'
            '  <Cat:Ctas>\n%s\n  </Cat:Ctas>\n'
            '</Cat:Catalogo>\n' % (esc(rfc), ejercicio, mes_s, hoy, ctas))

    def to_dict(self) -> List[Dict[str, Any]]:
        return [c.as_dict() for c in self.cuentas]
