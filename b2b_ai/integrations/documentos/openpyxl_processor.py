# -*- coding: utf-8 -*-
"""openpyxl_processor.py — Excel generation via openpyxl."""
from __future__ import annotations

import logging
import os
import time
from typing import Any, Dict

from b2b_ai.integrations.documentos.adapter import DocumentProcessor
from b2b_ai.integrations.documentos.models import (
    ExcelExportRequest, ExcelExportResult, OCRRequest, OCRResult,
    PDFRequest, PDFResult, XMLParseRequest, XMLParseResult,
)

logger = logging.getLogger(__name__)


class OpenPyxlProcessor(DocumentProcessor):
    """Excel processing using openpyxl."""

    def __init__(self):
        super().__init__(name="OpenPyxlProcessor")

    def generate_pdf(self, request: PDFRequest) -> PDFResult:
        raise NotImplementedError("OpenPyxlProcessor does not generate PDF.")

    def parse_xml(self, request: XMLParseRequest) -> XMLParseResult:
        raise NotImplementedError("OpenPyxlProcessor does not parse XML.")

    def extract_text_ocr(self, request: OCRRequest) -> OCRResult:
        raise NotImplementedError("OpenPyxlProcessor does not do OCR.")

    def export_excel(self, request: ExcelExportRequest) -> ExcelExportResult:
        start = time.time()
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = request.sheet_name
            headers = request.headers or (list(request.data[0].keys()) if request.data else [])
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            for row_idx, row_data in enumerate(request.data, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))
            if request.column_widths:
                for col_name, width in request.column_widths.items():
                    ws.column_dimensions[col_name].width = width
            output = request.output_path or "/tmp/openpyxl_output.xlsx"
            wb.save(output)
            size = os.path.getsize(output)
            elapsed = (time.time() - start) * 1000
            return ExcelExportResult(
                success=True, file_path=output, file_size=size,
                num_rows=len(request.data), num_columns=len(headers),
            )
        except ImportError:
            logger.warning("OpenPyxlProcessor: openpyxl not installed")
            return ExcelExportResult(success=False, file_path="")
        except Exception as e:
            logger.error(f"OpenPyxlProcessor: export failed: {e}")
            return ExcelExportResult(success=False, file_path="")
